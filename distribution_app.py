import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import plotly.graph_objects as go
from datetime import datetime
import io

# ════════════════════════════════════════════════════════════════
# 常數
# ════════════════════════════════════════════════════════════════

G1_STORES = ['彰草店', '金美店', '日華店']
G2_STORES = ['北屯店', '向上店', '五權店', '太平店', '大雅店', '漢口店']
G3_STORES = ['金馬店', '正德店', '大埔店', '三民店', '線東店', '彰美店',
             '過溝店', '彰鹿店', '泰和店', '精誠店', '秀二店', '花壇店', '華山店']

G1_ITEMS  = ['特幼', '幼大口', '多粒', '多大口', '幼菁', '雙子星']
G2_ITEMS  = ['特幼', '多菁',   '幼大口', '多粒', '多大口', '幼菁', '雙子星']
G3_ITEMS  = ['特幼', '普通',   '幼大口', '多粒', '多大口', '幼菁', '雙子星']

ITEM_COLS = {  # (品名欄, 售量欄) 1-indexed，與粒數累計 sheet 一致
    '特幼':   (2,  3),  '多菁':   (4,  5),  '普通':   (4,  5),
    '幼大口': (6,  7),  '多粒':   (8,  9),  '多大口': (10, 11),
    '幼菁':   (12, 13), '雙子星': (14, 15),
}

GROUPS = [
    ('日紅', G1_STORES, G1_ITEMS),
    ('台中', G2_STORES, G2_ITEMS),
    ('彰化', G3_STORES, G3_ITEMS),
]
ALL_STORES = G1_STORES + G2_STORES + G3_STORES

# ════════════════════════════════════════════════════════════════
# 讀取庫存（Raw_data 下班量）
# ════════════════════════════════════════════════════════════════

def load_inventory(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl')
    cs = ci = cq = None
    for col in df.columns:
        s = str(col)
        if '店名'  in s: cs = col
        if '品名'  in s: ci = col
        if '下班量' in s: cq = col
    if not all([cs, ci, cq]):
        return None, "找不到 店名／品名／下班量 欄位"
    df = df[[cs, ci, cq]].copy()
    df.columns = ['店名', '品名', '下班量']
    df['店名']  = df['店名'].astype(str).str.strip()
    df['品名']  = df['品名'].astype(str).str.strip()
    df = df[df['店名'].ne('') & df['店名'].ne('nan') & df['品名'].ne('nan')]
    df['下班量'] = pd.to_numeric(df['下班量'], errors='coerce').fillna(0).astype(int)
    inv = {}
    for _, row in df.iterrows():
        s, i, q = row['店名'], row['品名'], row['下班量']
        inv.setdefault(s, {})[i] = q
    return inv, None

# ════════════════════════════════════════════════════════════════
# 讀取粒數累計 xlsx 歷史售量
# ════════════════════════════════════════════════════════════════

def parse_date_key(name):
    """把 '4-3' 解析成可排序的 tuple (4, 3)"""
    try:
        parts = name.split('-')
        if len(parts) == 2:
            return (int(parts[0]), int(parts[1]))
    except:
        pass
    return None

def read_one_sheet(ws):
    """從單一 sheet 讀出每店每品項的售量"""
    store_set = set(ALL_STORES)
    data = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        store = str(row[0]).strip()
        if store not in store_set:
            continue
        sd = {}
        for item, (nc, qc) in ITEM_COLS.items():
            if len(row) <= qc - 1:
                continue
            name_val = row[nc - 1]
            qty_val  = row[qc - 1]
            if name_val == item and isinstance(qty_val, (int, float)):
                sd[item] = int(qty_val)
        if sd:
            data[store] = sd
    return data

def read_one_sheet_xlrd(ws_xlrd):
    """從 xlrd sheet 讀出每店每品項售量"""
    store_set = set(ALL_STORES)
    data = {}
    for ri in range(ws_xlrd.nrows):
        row = [ws_xlrd.cell_value(ri, ci) for ci in range(ws_xlrd.ncols)]
        if not row[0]: continue
        store = str(row[0]).strip()
        if store not in store_set: continue
        sd = {}
        for item, (nc, qc) in ITEM_COLS.items():
            if ws_xlrd.ncols <= qc - 1: continue
            name_val = row[nc - 1]
            qty_val  = row[qc - 1]
            if name_val == item and isinstance(qty_val, (int, float)) and qty_val != '':
                sd[item] = int(qty_val)
        if sd:
            data[store] = sd
    return data

def load_history(file_bytes, fname="", n_days=7):
    """讀取最近 n_days 個有效 sheet，支援 .xlsx 和 .xls"""
    history = {}
    try:
        if fname.lower().endswith(".xls"):
            import xlrd
            wb_x = xlrd.open_workbook(file_contents=file_bytes)
            dated = [(parse_date_key(n), n) for n in wb_x.sheet_names() if parse_date_key(n)]
            dated.sort(key=lambda x: x[0])
            for (m, d), name in dated[-n_days:]:
                history[f"{m}/{d}"] = read_one_sheet_xlrd(wb_x.sheet_by_name(name))
        else:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            dated = [(parse_date_key(n), n) for n in wb.sheetnames if parse_date_key(n)]
            dated.sort(key=lambda x: x[0])
            for (m, d), name in dated[-n_days:]:
                history[f"{m}/{d}"] = read_one_sheet(wb[name])
    except Exception as e:
        st.error(f"讀取累計檔失敗：{e}")
    return history

# ════════════════════════════════════════════════════════════════
# 配貨計算
# ════════════════════════════════════════════════════════════════

def proportional_alloc(total, weights):
    """整數按比例分配，確保加總等於 total"""
    total = int(total)
    if not weights or sum(weights) == 0:
        n = max(len(weights), 1)
        base = [total // n] * len(weights)
        for i in range(total - sum(base)):
            base[i] += 1
        return base
    s = sum(weights)
    floats = [total * w / s for w in weights]
    floors = [int(f) for f in floats]
    fracs  = sorted(enumerate(floats[i] - floors[i] for i in range(len(weights))),
                    key=lambda x: -x[1])
    shortage = total - sum(floors)
    for i in range(abs(shortage)):
        idx = fracs[i][0]
        floors[idx] += 1 if shortage > 0 else -1
    return floors

def calculate(inventory, history, n_avg=3):
    all_dates = sorted(history.keys())
    avg_dates = all_dates[-n_avg:]

    # 3日平均售量
    avg = {}
    for gname, stores, items in GROUPS:
        for store in stores:
            avg[store] = {}
            for item in items:
                daily = [history[d].get(store, {}).get(item, 0) for d in avg_dates]
                avg[store][item] = sum(daily) / len(daily) if daily else 0.0

    # 配貨差額（各區、各品項）
    dist = {s: {} for s in ALL_STORES}

    for gname, stores, items in GROUPS:
        for item in items:
            total_avail  = sum(inventory.get(s, {}).get(item, 0) for s in stores)
            store_avgs   = [avg[s].get(item, 0) for s in stores]
            allocations  = proportional_alloc(total_avail, store_avgs)
            for store, alloc in zip(stores, allocations):
                current = inventory.get(store, {}).get(item, 0)
                dist[store][item] = alloc - current

    return dist, avg

# ════════════════════════════════════════════════════════════════
# 建立顯示用 DataFrame
# ════════════════════════════════════════════════════════════════

def make_df(stores, items, values):
    """values: {store: {item: number}}，回傳 DataFrame（rows=items, cols=store縮寫）"""
    return pd.DataFrame(
        {s.replace('店', ''): {it: values.get(s, {}).get(it, 0) for it in items}
         for s in stores},
        index=items
    )

def style_dist(df):
    def _color(val):
        if pd.isna(val) or val == 0: return 'color: #aaa'
        return 'color: #d32f2f; font-weight:600' if val > 0 else 'color: #1565c0; font-weight:600'
    return df.style.map(_color).format(lambda v: f"+{int(v)}" if v > 0 else str(int(v)))

def style_plain(df):
    return df.style.format(lambda v: str(int(v)) if pd.notna(v) else '—')

# ════════════════════════════════════════════════════════════════
# 折線圖
# ════════════════════════════════════════════════════════════════

def make_chart(stores, item, history):
    dates = sorted(history.keys())
    fig = go.Figure()
    for store in stores:
        short = store.replace('店', '')
        y = [history[d].get(store, {}).get(item, 0) for d in dates]
        fig.add_trace(go.Scatter(x=dates, y=y, mode='lines+markers', name=short))
    fig.update_layout(
        title=f"{item} — 近 {len(dates)} 日銷量",
        xaxis_title='日期', yaxis_title='銷量（包）',
        height=320, margin=dict(t=40, b=30)
    )
    return fig

# ════════════════════════════════════════════════════════════════
# 匯出 Excel（基礎版，格式待調整）
# ════════════════════════════════════════════════════════════════

def export_excel(dist, avg, inventory):
    wb  = Workbook()
    del wb['Sheet']
    _s  = Side(style='thin')
    BRD = Border(left=_s, right=_s, top=_s, bottom=_s)
    AC  = Alignment(horizontal='center', vertical='center')

    def F(bold=False, c="000000"):
        return Font(name='新細明體', size=11, bold=bold, color=c)

    def sc(ws, r, c, v=None, f=None):
        cl = ws.cell(r, c)
        if v is not None: cl.value = v
        cl.font = f or F()
        cl.alignment = AC
        cl.border = BRD
        return cl

    for gname, stores, items in GROUPS:
        ws = wb.create_sheet(gname)
        short = [s.replace('店', '') for s in stores]

        # Header
        sc(ws, 1, 1, '品名', F(bold=True))
        for ci, s in enumerate(short, 2):
            sc(ws, 1, ci, s, F(bold=True))
        ws.column_dimensions['A'].width = 8
        for ci in range(2, len(stores) + 2):
            ws.column_dimensions[chr(64 + ci)].width = 6

        # Distribution rows
        for ri, item in enumerate(items, 2):
            sc(ws, ri, 1, item, F(bold=True))
            for ci, store in enumerate(stores, 2):
                val = dist.get(store, {}).get(item, 0)
                color = "D32F2F" if val > 0 else ("1565C0" if val < 0 else "888888")
                sc(ws, ri, ci, val if val != 0 else None, F(c=color))

        # Blank row then 參考資訊
        ref_row = len(items) + 3
        sc(ws, ref_row, 1, '3日均量', F(bold=True))
        for ci, store in enumerate(stores, 2):
            sc(ws, ref_row, ci, f"", F())
        for ri2, item in enumerate(items, ref_row + 1):
            sc(ws, ri2, 1, item, F())
            for ci, store in enumerate(stores, 2):
                sc(ws, ri2, ci, round(avg.get(store, {}).get(item, 0), 1), F())

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ════════════════════════════════════════════════════════════════
# Streamlit UI
# ════════════════════════════════════════════════════════════════

st.set_page_config(page_title="配貨系統", layout="wide")
st.title("🚚 配貨系統")

c1, c2 = st.columns(2)
with c1:
    f_raw = st.file_uploader("① 從檳榔管理系統匯出之原始檔（含下班量）", type=['xlsx'])
with c2:
    f_cum = st.file_uploader("② 粒數累計 .xlsx / .xls（讀取歷史售量）", type=['xlsx', 'xls'])

if not (f_raw and f_cum):
    st.info("請上傳兩份檔案以開始配貨計算。")
    st.stop()

# ── 載入資料 ──
inv, err = load_inventory(f_raw.getvalue())
if err:
    st.error(err); st.stop()

history = load_history(f_cum.getvalue(), fname=f_cum.name, n_days=7)
if not history:
    st.error("累計檔找不到有效的歷史工作表（格式需為 M-D，例如 4-3）"); st.stop()

n_avail = len(history)
dates_str = '、'.join(sorted(history.keys()))
st.caption(f"✅ 庫存：{len(inv)} 家店　｜　歷史：最近 {n_avail} 天（{dates_str}）　｜　計算基準：近 {min(3, n_avail)} 天平均 × 2")

dist, avg = calculate(inv, history, n_avg=min(3, n_avail))

# ── 分 tab 顯示各區 ──
tabs = st.tabs([f"📍 {g}" for g, _, _ in GROUPS])

for tab, (gname, stores, items) in zip(tabs, GROUPS):
    with tab:

        # 配貨差額表（可編輯）
        st.markdown("#### 配貨差額　🔴 補　🔵 退")
        dist_df = make_df(stores, items, dist)

        edited = st.data_editor(
            dist_df,
            use_container_width=True,
            key=f"edit_{gname}",
            column_config={
                col: st.column_config.NumberColumn(col, format="%+d")
                for col in dist_df.columns
            }
        )

        # 同步回 dist（讓匯出拿到手動修改後的值）
        for store in stores:
            short = store.replace('店', '')
            for item in items:
                if short in edited.columns and item in edited.index:
                    dist[store][item] = int(edited.loc[item, short])

        # 參考資訊（摺疊）
        with st.expander("📊 參考資訊（3日均量 ／ 現有庫存）"):
            rc1, rc2 = st.columns(2)
            with rc1:
                st.markdown("**3 日平均售量**")
                st.dataframe(style_plain(make_df(stores, items, avg)),
                             use_container_width=True)
            with rc2:
                st.markdown("**現有庫存（下班量）**")
                st.dataframe(style_plain(make_df(stores, items, inv)),
                             use_container_width=True)

        # 7日折線圖（摺疊，點選品項）
        with st.expander("📈 7 日銷量折線圖"):
            sel = st.selectbox("選擇品項", items, key=f"sel_{gname}")
            st.plotly_chart(make_chart(stores, sel, history),
                            use_container_width=True)

# ── 匯出 ──
st.markdown("---")
if st.button("📥 匯出配貨表 Excel", type="primary"):
    xlsx = export_excel(dist, avg, inv)
    st.download_button(
        "💾 下載配貨表",
        data=xlsx,
        file_name=f"配貨表_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
